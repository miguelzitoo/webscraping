from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
import time 
import openpyxl

# informação OAB e estado
numero_oab = input('numero OAB ')
estado = input('Estado: Ex SP, MG, AL.. ')

# Abrir o navegador (google)
driver = webdriver.Chrome(ChromeDriverManager().install())
# entrar no site PJE
driver.get('https://pje-consulta-publica.tjmg.jus.br/')
# inserir as informações
time.sleep(5)
campo_oab = driver.find_element(By.XPATH, "//*[@id='fPP:Decoration:numeroOAB']")
campo_oab.send_keys(numero_oab)
dropdown_estados = driver.find_element(By.XPATH, "//*[@id='fPP:Decoration:estadoComboOAB']")
opcoes_estados = Select(dropdown_estados)
# clicar em pesquisar
opcoes_estados.select_by_visible_text(estado)
driver.find_element(By.XPATH, '//*[@id="fPP:searchProcessos"]').click()
time.sleep(10)
# entrar em cada um dos processos
processos = driver.find_elements(By.XPATH, "//b[@class='btn-block']")

for processo in processos:
    processo.click()
    time.sleep(5)
    # identificar as janelas e alterar para a nova
    janelas = driver.window_handles
    driver.switch_to.window(janelas[-1])
    driver.set_window_size(1920, 1080)
    # pesquisar o numero de processos e selecionar pelo Xpath
    numero_processo = driver.find_elements(By.XPATH, "//div[@class='col-sm-12 ']")
    numero_processo = numero_processo[0]
    numero_processo = numero_processo.text
    # pesquisar a data do processo e selecionar pelo Xpath
    data_processo = driver.find_elements(By.XPATH, "//div[@class='value col-sm-12 ']")
    data_processo = data_processo[1]
    data_processo = data_processo.text
    # 
    movimentacoes = driver.find_elements(By.XPATH, "//div[@id='j_id132:processoEventoPanel']//tr[contains(@class,'rich-table-row')]//td//div//span")
    lista_movimentacoes = []
    for movimentacao in movimentacoes:
        lista_movimentacoes.append(movimentacao.text)

    workbook = openpyxl.load_workbook("automation\dados.xlsx")
    try:
        pagina_processo = workbook[numero_processo]
        pagina_processo['A1'].value = "Numero Processo"
        pagina_processo['B1'].value = "Data distribuição"
        pagina_processo['C1'].value = "Movimentações"
        pagina_processo['A2'].value = numero_processo
        pagina_processo['B2'].value = data_processo
        for index, linha in enumerate(pagina_processo.iter_rows(min_row=2,max_row = len(lista_movimentacoes), min_col = 3, max_col = 3)):
            for celula in linha:
                celula.value = lista_movimentacoes[index]
        workbook.save('dados.xlsx')
        driver.close()
        time.sleep(5)
        driver.switch_to.window(driver.window_handles[0])
    except Exception as error:
        workbook.create_sheet(numero_processo)
        pagina_processo = workbook[numero_processo] 
        pagina_processo['A1'].value = "Numero Processo"
        pagina_processo['B1'].value = "Data distribuição"
        pagina_processo['C1'].value = "Movimentações"
        pagina_processo['A2'].value = numero_processo
        pagina_processo['B2'].value = data_processo
        for index, linha in enumerate(pagina_processo.iter_rows(min_row=2,max_row = len(lista_movimentacoes), min_col = 3, max_col = 3)):
            for celula in linha:
                celula.value = lista_movimentacoes[index]
        workbook.save('dados.xlsx')
        driver.close()
        time.sleep(5)
        driver.switch_to.window(driver.window_handles[0])

